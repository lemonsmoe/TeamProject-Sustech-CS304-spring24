import openpyxl
from openpyxl.styles import PatternFill, Alignment

all_color = ['8db5e8', 'ffb47f', 'a2d4a2', 'e88b8b', 'c9b5e8', 'b8a39e', 'f2b5cc', 'bcbcbc', 'd9da8b', '7bd1e8',
                     'c6e9f0', '4caf50']
# 列映射，将数字星期映射为Excel表的列数
WEEKDAYS_COLUMN_MAPPING = {
    1: 'B', # 星期一
    2: 'C', # 星期二
    3: 'D', # 星期三
    4: 'E', # 星期四
    5: 'F', # 星期五
    6: 'G', # 星期六
    7: 'H', # 星期日
}

def write_schedule_scheme_to_excel(scheduler, output_excel_path: str):
    import warnings

    # 在你的代码之前添加以下代码
    warnings.filterwarnings("ignore", category=FutureWarning)
    # 创建Excel工作薄
    workbook = openpyxl.Workbook()
    # 删除默认创建的sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # 循环排课方案
    for index, courses in enumerate(scheduler.schedule_scheme):
        # 为每个排课方案创建一个新的sheet
        sheet = workbook.create_sheet(title=f'Scheme {index + 1}')

        # 写入列（星期）header
        for day, column in WEEKDAYS_COLUMN_MAPPING.items():
            sheet[f'{column}1'] = f'星期{day}'

        # 写入行（节次）header
        for period in range(1, 7):
            sheet[f'A{period + 1}'] = f'第{period}节'

        # 准备一个颜色迭代器
        color_iter = iter(all_color)

        # 用来记录哪些课程已经分配了颜色
        course_color_mapping = {}

        # 设置列宽
        for column in list(WEEKDAYS_COLUMN_MAPPING.values()):
            sheet.column_dimensions[column].width = 20

        # 设置行高
        for row in range(2, 8):  # 假设我们有6节课，所在行号从2开始
            sheet.row_dimensions[row].height = 70

        # 计算学分
        credit = 0
        # 遍历这个方案里的所有课程
        for course in courses:
            credit += float(course.all_info[7])

            # 如果还没给这门课程分配颜色，则分配一个新的颜色
            if course.name not in course_color_mapping:
                course_color_mapping[course.name] = next(color_iter)

            # 为课程的每个时间段填充颜色
            for time in course.time:
                weekday, start_period, end_period, odd_week, even_week = time
                # 计算Excel中对应的列
                column = WEEKDAYS_COLUMN_MAPPING[weekday]
                # 对于每一节课
                for period in range(start_period, end_period + 1):
                    cell = f'{column}{period + 1}'  # Excel中的单元格位置
                    # 设置单元格内容
                    # 如果这个单元格已经有内容了，课程信息就直接追加在后面
                    if sheet[cell].value:
                        sheet[cell].value += f'\n{course.name}' if odd_week and even_week else f'\n{course.name}(单)' if odd_week else f'\n{course.name}(双)'
                        sheet[cell].value += '\n' + course.all_info[12]
                    else:
                        sheet[cell] = course.name if odd_week and even_week else course.name + '(单)' if odd_week else course.name + '(双)'
                        sheet[cell].value += '\n'+course.all_info[12]
                    # 设置填充颜色
                    fill = PatternFill(start_color=course_color_mapping[course.name],
                                       end_color=course_color_mapping[course.name],
                                       fill_type='solid')
                    sheet[cell].fill = fill

        # 在A1单元格写入学分
        sheet['A1'] = f'学分:\n{credit}'
        # 在B9单元格写入所有课程的名称
        sheet['B9'] = '该方案选择的课:  ' + '    '.join([course.name for course in courses])
        # 在B11单元格写入exclude_time
        b11_str = '已排除的时间段:  '
        for key in scheduler.exclude_time:
            b11_str += f'{key}: {scheduler.exclude_time[key]}\n'
        sheet['B11'] = b11_str


        # 设置文本居中
        center_aligned_text = Alignment(wrap_text=True, horizontal='center', vertical='center')
        # 为每个表格中的单元格（包括头部）设置居中
        for row in sheet.iter_rows(min_row=1, max_col=8, max_row=7):  # 假设共7行 (包括头部), 8列
            for cell in row:
                cell.alignment = center_aligned_text

    # 保存工作薄到指定的路径
    workbook.save(output_excel_path)